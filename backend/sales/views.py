from rest_framework import generics
from .serializers import (
    SaleSerializer,
    CheckoutSerializer,
    SaleReturnSerializer,
)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny

from products.models import Product
from inventory.models import Inventory, StockMovement

from .models import Sale, SaleItem, SaleReturn

from accounts.models import User

from django.shortcuts import get_object_or_404

from django.db.models import Sum, Avg
from django.utils import timezone
from django.db.models import Q

from django.db import transaction

from notifications.utils import create_notification
from notifications.models import Notification
from rest_framework.decorators import api_view, permission_classes
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from audit.utils import create_audit_log
from audit.models import AuditLog




class SaleListCreateView(generics.ListCreateAPIView):
    serializer_class = SaleSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = Sale.objects.all().order_by("-created_at")

        receipt = self.request.query_params.get("receipt")
        date = self.request.query_params.get("date")

        if receipt:
            queryset = queryset.filter(
                receipt_number__icontains=receipt
            )

        if date:
            queryset = queryset.filter(
                created_at__date=date
            )

        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        serializer = self.get_serializer(
            queryset,
            many=True
        )

        completed_sales = queryset.filter(
            status=Sale.Status.COMPLETED
        )

        revenue = (
            completed_sales.aggregate(
                total=Sum("total_amount")
            )["total"] or 0
        )

        average_sale = (
            completed_sales.aggregate(
                avg=Avg("total_amount")
            )["avg"] or 0
        )

        today_sales = completed_sales.filter(
            created_at__date=timezone.now().date()
        ).count()

        return Response({
            "summary": {
                "total_revenue": revenue,
                "total_sales": completed_sales.count(),
                "average_sale": average_sale,
                "today_sales": today_sales,
            },
            "sales": serializer.data,
        })
class SaleDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sale.objects.all()
    serializer_class = SaleSerializer
    permission_classes = [AllowAny]

class CheckoutView(APIView):
    permission_classes = [AllowAny]

    @transaction.atomic
    def post(self, request):

        

        serializer = CheckoutSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data

        receipt_number = f"SALE-{timezone.now().strftime('%Y%m%d%H%M%S')}"

        cashier = User.objects.first()

        sale = Sale.objects.create(
            receipt_number=receipt_number,
            cashier=cashier,
            payment_method=data["payment_method"],
            discount=data["discount"],
            status=Sale.Status.COMPLETED,
        )

        for item in data["items"]:

            product = get_object_or_404( Product, pk=item["product"]
            )

            inventory = get_object_or_404( Inventory.objects.select_for_update(), product=product
            )

            if inventory.quantity < item["quantity"]:
                return Response(
                    {
                        "error": f"Not enough stock for {product.name}"
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            SaleItem.objects.create(
                sale=sale,
                product=product,
                quantity=item["quantity"],
                unit_price=product.selling_price,
            )

            inventory.quantity -= item["quantity"]
            inventory.save()

            StockMovement.objects.create(
                inventory=inventory,
                movement_type=StockMovement.MovementType.SALE,
                quantity=item["quantity"],
                note=f"Sale {sale.receipt_number}",
            )

        sale.calculate_totals()
        sale.save()

        create_audit_log(
            user=request.user,
            module="Sales",
            action=AuditLog.Action.CREATE,
            description=(
                f"Completed sale {sale.receipt_number} "
                f"worth ₦{sale.total_amount}"
            ),
            object_id=sale.id,
        )

        create_notification(
            user=request.user,
            title="New Sale",
            message=f"Receipt {sale.receipt_number} completed successfully.",
            notification_type=Notification.NotificationType.SALE,
        )

        return Response(
            {
                "message": "Sale completed successfully.",
                "receipt_number": sale.receipt_number,
            },
            status=status.HTTP_201_CREATED,
        )

class ReturnSaleAPIView(APIView):
    permission_classes = [AllowAny]

    @transaction.atomic
    def post(self, request, sale_id):
        sale = get_object_or_404(
            Sale,
            id=sale_id,
            status=Sale.Status.COMPLETED
        )

        if hasattr(sale, "sale_return"):
            return Response(
                {
                    "detail": "This sale has already been returned."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        reason = request.data.get("reason", "")

        for item in sale.items.all():

            inventory = Inventory.objects.select_for_update().get(
                product=item.product
            )

            inventory.quantity += item.quantity
            inventory.save()

            StockMovement.objects.create(
                inventory=inventory,
                movement_type=StockMovement.MovementType.RETURN,
                quantity=item.quantity,
                note=f"Returned Sale {sale.receipt_number}",
            )

        sale.status = Sale.Status.RETURNED
        sale.save(update_fields=["status"])

        create_audit_log(
            user=request.user,
            module="Sales",
            action=AuditLog.Action.UPDATE,
            description=(
                f"Returned sale {sale.receipt_number}"
            ),
            object_id=sale.id,
        )

        sale_return = SaleReturn.objects.create(
            sale=sale,
            reason=reason,
            refund_amount=sale.total_amount,
            returned_by=User.objects.first(),
        )

        serializer = SaleReturnSerializer(sale_return)

        return Response(
            {
                "message": "Sale returned successfully.",
                "return": serializer.data,
            },
            status=status.HTTP_201_CREATED,
        )

class SaleReturnListAPIView(generics.ListAPIView):
    serializer_class = SaleReturnSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = (
            SaleReturn.objects
            .select_related("sale", "returned_by")
            .order_by("-returned_at")
        )

        receipt = self.request.query_params.get("receipt")
        date = self.request.query_params.get("date")

        if receipt:
            queryset = queryset.filter(
                sale__receipt_number__icontains=receipt
            )

        if date:
            queryset = queryset.filter(
                returned_at__date=date
            )

        return queryset

@api_view(["GET"])
@permission_classes([AllowAny])
def SaleReceiptAPIView(request, sale_id):

    sale = get_object_or_404(
        Sale.objects.prefetch_related("items__product"),
        id=sale_id,
    )

    serializer = SaleSerializer(sale)

    return Response(serializer.data)
@api_view(["GET"])
@permission_classes([AllowAny])
def ExportSalesCSVAPIView(request):

    receipt = request.GET.get("receipt")
    date = request.GET.get("date")

    sales = Sale.objects.all().order_by("-created_at")

    if receipt:
        sales = sales.filter(
            receipt_number__icontains=receipt
        )

    if date:
        sales = sales.filter(
            created_at__date=date
        )

    response = HttpResponse(content_type="text/csv")

    response["Content-Disposition"] = (
        'attachment; filename="sales.csv"'
    )

    writer = csv.writer(response)

    writer.writerow([
        "Receipt",
        "Date",
        "Cashier",
        "Payment",
        "Status",
        "Subtotal",
        "Discount",
        "Total",
    ])

    for sale in sales:
        writer.writerow([
            sale.receipt_number,
            sale.created_at.strftime("%d/%m/%Y %H:%M"),
            sale.cashier.username,
            sale.get_payment_method_display(),
            sale.get_status_display(),
            sale.subtotal,
            sale.discount,
            sale.total_amount,
        ])

    return response

@api_view(["GET"])
@permission_classes([AllowAny])
def ExportSalesExcelAPIView(request):

    receipt = request.GET.get("receipt")
    date = request.GET.get("date")

    sales = Sale.objects.all().order_by("-created_at")

    if receipt:
        sales = sales.filter(
            receipt_number__icontains=receipt
        )

    if date:
        sales = sales.filter(
            created_at__date=date
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"

    headers = [
        "Receipt",
        "Date",
        "Cashier",
        "Payment",
        "Status",
        "Subtotal",
        "Discount",
        "Total",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    row = 2

    for sale in sales:
        sheet.cell(row=row, column=1).value = sale.receipt_number
        sheet.cell(row=row, column=2).value = sale.created_at.strftime("%d/%m/%Y %H:%M")
        sheet.cell(row=row, column=3).value = sale.cashier.username
        sheet.cell(row=row, column=4).value = sale.get_payment_method_display()
        sheet.cell(row=row, column=5).value = sale.get_status_display()
        sheet.cell(row=row, column=6).value = float(sale.subtotal)
        sheet.cell(row=row, column=7).value = float(sale.discount)
        sheet.cell(row=row, column=8).value = float(sale.total_amount)

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="sales.xlsx"'
    )

    workbook.save(response)

    return response
